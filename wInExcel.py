'''
书名、作者、简介、评分、评论人数、字数、分类、价格
'''
import xlwt
import openpyxl
title = ['书名','作者','简介','评分','评论人数','字数','分类','价格']
list = [['坏小孩','紫金陈','结婚第四年...','9.8','1000','18万','推理悬疑','9.9'],['坏小孩','紫金陈','结婚第四年...','9.8','1000','18万','推理悬疑','9.9'],['坏小孩','紫金陈','结婚第四年...','9.8','1000','18万','推理悬疑','9.9']]
#


# for i in range(len(title)):
#     sht1.write(0,i,title[i])
#     sht1.write(1,i,list[i])
# xls.save('./mydata.xls')
def wIn(list):
    xls = xlwt.Workbook()
    sht1 = xls.add_sheet('Sheet1')
    title = ['书名', '作者', '价格','出版社','出版时间']
    #创建标题
    for i in range(len(title)):
        sht1.write(0, i, title[i])
    #写入数据
    for i in range(len(list)):
        for j in range(len(list[i])):
            sht1.write(i+1,j,list[i][j])
    xls.save('./data.xls')
# wIn(list,title)
def openxl(list):
    title = ['书名', '作者', '简介', '评分', '评论人数', '字数', '分类', '价格']
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=0, title='oxl_sheet')
    for i in range(len(title)):
        ws.cell(row = 1,column = i+1).value = title[i]

    for i in range(len(list)):
        for j in range(len(list[i])):
            ws.cell(row = i+2,column = j+1).value = list[i][j]
    wb.save('./data.xlsx')
